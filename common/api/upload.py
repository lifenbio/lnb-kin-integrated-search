"""
파일 업로드 API
- 키워드 파일 업로드
- URL 파일 업로드
"""
from openpyxl import load_workbook
from ninja import Router, File
from ninja.files import UploadedFile
from common.models import Keyword, URL


upload_router = Router()


@upload_router.post(
    "/keyword",
    auth=None
)
def upload_keyword(request, file: UploadedFile = File(None)):
    """키워드 엑셀 파일 업로드"""
    Keyword.objects.all().delete()

    wb = load_workbook(file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        Keyword.objects.get_or_create(
            product_name=(row[0] or '') if len(row) > 0 else '',
            keyword=(row[1] or '') if len(row) > 1 else '',
            priority=(row[2] or '') if len(row) > 2 else ''
        )

    return {"message": "키워드 파일 업로드 완료."}


@upload_router.post(
    "/url",
    auth=None
)
def upload_url(request, file: UploadedFile = File(None)):
    """URL 엑셀 파일 업로드"""
    URL.objects.all().delete()

    wb = load_workbook(file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        URL.objects.get_or_create(
            url=(row[0] or '') if len(row) > 0 else '',
            product_name=(row[1] or '') if len(row) > 1 else '',
            conversion_keyword=(row[2] or '') if len(row) > 2 else '',
            content_type=(row[3] or '') if len(row) > 3 else '',
            keyword=(row[4] or '') if len(row) > 4 else ''
        )

    return {"message": "URL 파일 업로드 완료."}
